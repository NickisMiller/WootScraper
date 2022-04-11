#https://www.profitguru.com/calculator/fba
#https://sellercentral.amazon.com/revcal?ref=RC1


from asyncio.windows_events import NULL
from calendar import c
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
import time
import openpyxl


#get to asin site
driver = webdriver.Chrome()
driver.get("https://www.profitguru.com/calculator/fba")
time.sleep(1)


#get the asin column from sheet
asin_and_price = pd.read_excel(r"C:\Users\nickm\OneDrive\Desktop\python\demo.xlsx")
#price = pd.read_excel(r"C:\Users\nickm\OneDrive\Desktop\python\demo.xlsx", usecols="B")

time.sleep(1)
asin_input = driver.find_element_by_id("load_asin_asin")
cost_price_input = driver.find_element_by_id("fba_calculation_supplierPrice")
search_button = driver.find_element_by_id("search-button")

size = len(asin_and_price)
for i in range(0,size):
    asin_number = asin_and_price['asin'].loc[i]
    excel_price = asin_and_price['price'].loc[i]
    #need to remove the dollar sign
    price = excel_price[1:]
    time.sleep(2)
    asin_input.clear()
    time.sleep(2)
    asin_input.send_keys(asin_number)
    time.sleep(1)

    search_button.click()
    time.sleep(1)
    cost_price_input.clear()
    cost_price_input.send_keys(price)
    time.sleep(1)
    calculate_button = driver.find_element_by_xpath("/html/body/main/div/div[3]/div[1]/div[2]/form/div[2]/div[1]/button")
    calculate_button.click()
    time.sleep(4)

    net_profit = driver.find_element_by_xpath("/html/body/main/div/div[3]/div[1]/div[2]/form/div[1]/div[1]/div/div/div/table/tbody/tr[7]/td[1]/span").text
    roi = driver.find_element_by_xpath("/html/body/main/div/div[3]/div[1]/div[2]/form/div[1]/div[1]/div/div/div/table/tbody/tr[9]/td[1]/span").text
    amazon_price = driver.find_element_by_xpath("/html/body/main/div/div[3]/div[1]/div[2]/div/div[2]/div/div/div/table/tbody/tr[1]/td[2]/div[1]/span[2]").text

    print(net_profit)
    print(roi)
    print(amazon_price)
    print(i)
  

    xfile = openpyxl.load_workbook('demo.xlsx')

    sheet = xfile.get_sheet_by_name('Sheet1')
    
    x=i+2
    
    #sheet["D"] = amazon_price
    sheet.cell(row = x, column = 4).value = amazon_price
    sheet.cell(row = x, column = 5).value = net_profit
    sheet.cell(row = x, column = 6).value = roi

    xfile.save('demo.xlsx')

    time.sleep(3)
    

    



########### TODO ###############
# READ CSV ----- DONE
# INPUT ASIN AND PRICE INTO CALCULATOR ---- DONE 
# GRAB IMPORTANT INFO FROM CALCULATOR (PROFIT, ROI, AMAZON PRICE)